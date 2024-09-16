from flask import Flask, request, send_file, abort, redirect, render_template
import pandas as pd
import io
import re
from openpyxl.styles import PatternFill

app = Flask(__name__)

def clean_numeric(value):
    """Remove non-numeric characters from a string."""
    return float(re.sub(r'[^\d.,-]', '', str(value)).replace(',', '.'))

from datetime import datetime, timedelta

def check_criteria(row, date=None):
    """Check if a coworker meets the criteria for a raise."""
    absentismo_ok = row['Absentismo'] <= 8 
    disciplina_ok = row['Processos Disciplinares'] == 0
    date_ok = True
    if date:
        if 'Ultima Revisao' in row.index:
            last_review_date = row['Ultima Revisao']
            if isinstance(last_review_date, str):
                last_review_date = datetime.strptime(last_review_date, '%Y-%m-%d')  # assuming the date is in 'YYYY-MM-DD' format
            date_ok = (datetime.strptime(date, '%Y-%m-%d') - last_review_date).days >= 365
        else:
            # Column 'Ultima Revisao' is missing; default to passing the date check
            date_ok = True
    return absentismo_ok and disciplina_ok and date_ok


def aplicacao_ajuste(excel_data, general_raise, group_raise=None, group=None, date=None):
    # Apply filters
    excel_data['Eligible for Raise'] = excel_data.apply(check_criteria, args=(date,), axis=1)
    
    if group:
        # Apply group-specific raise
        group_mask = excel_data['Eligible for Raise'] & (excel_data['Designação CategProf'] == group)
        excel_data.loc[group_mask, ['Vencimento+IHT', 'Vencimento', 'Isenção de Horário']] *= (1 + group_raise)
        
        # Apply general raise to all except the specific group
        general_mask = excel_data['Eligible for Raise'] & ~group_mask
        excel_data.loc[general_mask, ['Vencimento+IHT', 'Vencimento', 'Isenção de Horário']] *= (1 + general_raise)
        
    else:
        # Apply general raise to all eligible employees
        excel_data.loc[excel_data['Eligible for Raise'], ['Vencimento+IHT', 'Vencimento', 'Isenção de Horário']] *= (1 + general_raise)
    
    failed_criteria = ~excel_data['Eligible for Raise']
    return excel_data, failed_criteria


def compare_and_highlight(excel_data1, excel_data2):
    highlight_indices = []

    for index, row in excel_data1.iterrows():
        categoria = row['Designação CategProf']
        vencimento_iht = row['Vencimento+IHT']
        
        matching_row = excel_data2[excel_data2['CATEGORIA PROFISSIONAL'] == categoria]
        
        if not matching_row.empty:
            limite_superior = clean_numeric(matching_row['LIMITE SUPERIOR'].values[0])
            limite_inferior = clean_numeric(matching_row['LIMITE INFERIOR'].values[0])
            
            if vencimento_iht >= limite_superior or vencimento_iht <= limite_inferior:
                highlight_indices.append(index)
    
    return highlight_indices

@app.route('/')
def index():
    return render_template('index.html', title='Home')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file1' not in request.files or request.files['file1'].filename == '':
            return 'Main file must be uploaded'
        
        file1 = request.files['file1']
        group = request.form.get('groupselect')
        date = request.form.get('date')
        general_percentage = request.form.get('percentage')
        group_percentage = request.form.get('group_percentage')
        
        if general_percentage is None or general_percentage.strip() == '':
            general_percentage = 0
        else:
            general_percentage = float(general_percentage) / 100
        
        if group_percentage and group_percentage.strip() != '':
            group_percentage = float(group_percentage) / 100
        else:
            group_percentage = 0
        
        df1 = pd.read_excel(file1)
        
        # Check if 'Ultima Revisao' column exists; if not, handle accordingly
        if 'Ultima Revisao' in df1.columns:
            df1['Ultima Revisao'] = pd.to_datetime(df1['Ultima Revisao'], errors='coerce')
        else:
            # If the column is missing, fill with NaT (Not a Time) for comparison
            df1['Ultima Revisao'] = pd.NaT
        
        necessary_columns_file1 = ['Designação CategProf', 'Vencimento+IHT', 'Vencimento', 'Isenção de Horário', 'Absentismo', 'Processos Disciplinares']
        for column in necessary_columns_file1:
            if column not in df1.columns:
                return f"Column '{column}' is missing in the main file (file1)."
        df1['Absentismo'] = pd.to_numeric(df1['Absentismo'], errors='coerce').fillna(0)
        df1['Processos Disciplinares'] = pd.to_numeric(df1['Processos Disciplinares'], errors='coerce').fillna(0)
        original_df1 = df1.copy()
        
        df1, failed_criteria = aplicacao_ajuste(df1, general_percentage, group_raise=group_percentage, group=group, date=date)
        
        comparison_html = ''
        warning_message = ''
        highlight_indices = []
        if 'file2' in request.files and request.files['file2'].filename != '':
            file2 = request.files['file2']
            df2 = pd.read_excel(file2, header=1)

            necessary_columns_file2 = ['CATEGORIA PROFISSIONAL', 'LIMITE SUPERIOR', 'LIMITE INFERIOR']
            for column in necessary_columns_file2:
                if column not in df2.columns:
                    return f"Column '{column}' is missing in the comparison file (file2)."
            
            highlight_indices = compare_and_highlight(df1, df2)
            comparison_html = create_comparison_html(original_df1, df1, highlight_indices, failed_criteria)
        else:
            warning_message = '<p>No comparison file uploaded. Proceeding without comparison.</p>'
            comparison_html = create_comparison_html(original_df1, df1, [], failed_criteria)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Highlight failed criteria in red and salary bands in yellow
            for sheet_name, data, criteria in [
                ('Processed Data', df1, failed_criteria),
                ('Failed Requirements', original_df1[failed_criteria], None)
            ]:
                data.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                for i, row in data.iterrows():
                    row_num = i + 2  # +2 to account for header row and 0-indexing
                    fill = None
                    if sheet_name != 'Failed Requirements' and failed_criteria.iloc[i]:
                        fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Red fill for failed criteria
                    elif sheet_name == 'Processed Data' and i in highlight_indices:
                        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Yellow fill for salary band issues
                    if fill:
                        for col in range(len(data.columns)):
                            worksheet.cell(row=row_num, column=col+1).fill = fill
        
        output.seek(0)
        
        return f'''
        <html>
            <head>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 20px;
                    }}
                    h1 {{
                        color: #333;
                    }}
                    table {{
                        border-collapse: collapse;
                        width: 100%;
                        margin-bottom: 20px;
                    }}
                    table, th, td {{
                        border: 1px solid #ccc;
                    }}
                    th, td {{
                        padding: 8px;
                        text-align: left;
                    }}
                    th {{
                        background-color: #f2f2f2;
                    }}
                    tr:nth-child(even) {{
                        background-color: #f9f9f9;
                    }}
                    tr:hover {{
                        background-color: #f1f1f1;
                    }}
                    .warning {{
                        color: #d9534f;
                        font-weight: bold;
                    }}
                    .success {{
                        color: #5cb85c;
                        font-weight: bold;
                    }}
                    .download-button {{
                        display: inline-block;
                        padding: 10px 20px;
                        font-size: 16px;
                        color: white;
                        background-color: #5cb85c;
                        border: none;
                        border-radius: 5px;
                        cursor: pointer;
                        text-decoration: none;
                    }}
                    .download-button:hover {{
                        background-color: #4cae4c;
                    }}
                </style>
            </head>
            <body>
                <h1>Processed Data</h1>
                {warning_message}
                {comparison_html}
                <form action="/download" method="post">
                    <input type="hidden" name="data" value="{output.getvalue().hex()}" />
                    <input type="submit" class="download-button" value="Download Processed Data" />
                </form>
            </body>
        </html>
        '''
    
    except Exception as e:
        return f"An error occurred: {str(e)}"

@app.route('/download', methods=['POST'])
def download_file():
    data = request.form.get('data')
    if data is None:
        abort(400, description="No data provided for download.")
    data = bytes.fromhex(data)
    return send_file(io.BytesIO(data), as_attachment=True, download_name='processed_data.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def create_comparison_html(original_df, processed_df, highlight_indices, failed_criteria):
    
    # Comparing 'Vencimento+IHT' only
    comparison_field = 'Vencimento+IHT'
    
    # Fields to display
    columns_to_display = ['Nome', 'Nº', 'Designação CategProf', 'Vencimento+IHT', 'Ultima Revisao', 'Absentismo', 'Processos Disciplinares']
    
    comparison_html = '''
    <style>
        .comparison-container {
            border: 1px solid #ccc;
            border-radius: 5px;
            padding: 10px;
            background-color: #f9f9f9;
            margin-bottom: 20px;
        }
        .comparison-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        .comparison-table th, .comparison-table td {
            padding: 10px;
            border: 1px solid #ddd;
        }
        .comparison-table th {
            background-color: #f4f4f4;
        }
        .comparison-table tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .comparison-table tr:hover {
            background-color: #e0e0e0;
        }
        .highlight {
            background-color: #ffff99;  /* Highlight color for important rows */
        }
        .failed {
            background-color: #f8d7da;  /* Red background for failed criteria */
        }
        .divider {
            height: 2px;
            background-color: #333;
            margin: 20px 0;
        }
        .employee-header {
            font-weight: bold;
            margin-bottom: 10px;
        }
        .go-to-bottom {
            display: inline-block;
            padding: 10px 20px;
            font-size: 16px;
            color: white;
            background-color: #007bff;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            text-decoration: none;
            position: fixed;
            bottom: 20px;
            right: 20px;
        }
        .go-to-bottom:hover {
            background-color: #0056b3;
        }
    </style>
    '''

    # Add a container for comparison and details
    comparison_html += '<div class="comparison-container">'
    
    for index, (orig_row, proc_row) in enumerate(zip(original_df.iterrows(), processed_df.iterrows())):
        _, orig_row = orig_row
        _, proc_row = proc_row
        
        row_style = ''
        if failed_criteria.iloc[index]:
            row_style = 'failed'
        elif index in highlight_indices:
            row_style = 'highlight'
        
        # Formatting the date to display only the date part
        ultima_revisao_date = orig_row['Ultima Revisao'].strftime('%Y-%m-%d') if pd.notna(orig_row['Ultima Revisao']) else 'N/A'
        
        # Adding comparison section for each employee
        comparison_html += f'''
        <div class="employee-header">
            {orig_row['Nome']} - Nº {orig_row['Nº']} - Última Revisão: {ultima_revisao_date}
        </div>
        <table class="comparison-table">
            <thead>
                <tr>
                    <th>Field</th>
                    <th>Original Value</th>
                    <th>Processed Value</th>
                    <th>Difference</th>
                </tr>
            </thead>
            <tbody>
                <tr class="{row_style}">
                    <td>Vencimento+IHT</td>
                    <td>{orig_row[comparison_field]}</td>
                    <td>{proc_row[comparison_field]}</td>
                    <td>{proc_row[comparison_field] - orig_row[comparison_field]}</td>
                </tr>
                <tr>
                    <td>Absentismo</td>
                    <td>{orig_row['Absentismo']}</td>
                    <td>{proc_row['Absentismo']}</td>
                    <td>{proc_row['Absentismo'] - orig_row['Absentismo']}</td>
                </tr>
                <tr>
                    <td>Processos Disciplinares</td>
                    <td>{orig_row['Processos Disciplinares']}</td>
                    <td>{proc_row['Processos Disciplinares']}</td>
                    <td>{proc_row['Processos Disciplinares'] - orig_row['Processos Disciplinares']}</td>
                </tr>
            </tbody>
        </table>
        <div class="divider"></div>
        '''
    
    comparison_html += '</div>'
    
    # Add "Go to the Bottom" button
    comparison_html += '''
    <a href="#" class="go-to-bottom" onclick="window.scrollTo(0, document.body.scrollHeight); return false;">
        Go to Downloads
    </a>
    '''

    return comparison_html






if __name__ == '__main__':
    app.run(debug=True)
